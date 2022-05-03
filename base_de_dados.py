from openpyxl import *


# Verificando as planilhas
def selecionar_planilha():
    try:
        dados = load_workbook('dados.xlsx')
    except:
        dados = Workbook()
        dados.save('dados.xlsx')
    else:
        dados = load_workbook('dados.xlsx')
        dados.save('dados.xlsx')
    return dados


def localizar_pagina_principal(planilha):
    pag_principal = planilha['Sheet']
    return pag_principal


# Salvando dados de login
def salvar_dados(planilha, pagina_principal, nome_de_usuario, senha):
    print('Usuario criado com sucesso.')
    pagina_principal.append([nome_de_usuario, str(senha)])
    planilha.save('dados.xlsx')


# Lendo dados da planilha
def ler_dados(pagina, nome_de_usuario, senha):
    verificacao = 0
    for linhas in pagina.iter_rows(min_row=0, max_row=100):
        if verificacao == 2:
            break
        verificacao = 0
        for cell in linhas:
            if nome_de_usuario == cell.value or senha == cell.value:
                verificacao += 1
                if verificacao == 2:
                    break
    if verificacao == 2:
        return True
    else:
        return False
